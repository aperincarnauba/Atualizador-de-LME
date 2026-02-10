import streamlit as st
import pandas as pd
import yfinance as yf
import requests
from bs4 import BeautifulSoup
from datetime import datetime, date
from openpyxl import load_workbook
import io

# --- CONFIGURAÇÕES ---
NOME_ABA = '2021_Base RPA'
DATA_HOJE = datetime.now().replace(hour=0, minute=0, second=0, microsecond=0)

month_map = {
    'jan': 1, 'fev': 2, 'mar': 3, 'abr': 4, 'mai': 5, 'jun': 6,
    'jul': 7, 'ago': 8, 'set': 9, 'out': 10, 'nov': 11, 'dez': 12
}

# --- FUNÇÕES AUXILIARES ---

def clean_val(v, tipo='metal'):
    if v is None:
        return None
    s = str(v).strip().upper()
    if 'FERIADO' in s or s in ['-', '']:
        return 'FERIADO'
    try:
        if tipo == 'metal':        # 2,604.00
            val = s.replace(',', '')
        else:                      # 5,40
            val = s.replace('.', '').replace(',', '.')
        return float(val)
    except:
        return s

@st.cache_data(ttl=3600)
def get_shock_metais(ano_inicio):
    data_rows = []
    hoje = date.today()
    headers = {'User-Agent': 'Mozilla/5.0'}

    for ano in range(ano_inicio, hoje.year + 1):
        for mes in range(1, 13):
            if ano == hoje.year and mes > hoje.month:
                break

            url = f'https://shockmetais.com.br/lme/{mes}-{ano}'
            try:
                res = requests.get(url, timeout=10, headers=headers, verify=False)
                if res.status_code != 200:
                    continue

                soup = BeautifulSoup(res.content, "html.parser")
                tbody = soup.find('tbody')
                if not tbody:
                    continue

                for r in tbody.find_all('tr'):
                    tds = r.find_all('td')
                    if len(tds) < 8 or 'Média' in tds[0].text:
                        continue

                    raw_date = tds[0].get_text(strip=True).lower()
                    if '/' not in raw_date:
                        continue

                    dia, m_abbr = raw_date.split('/')
                    m_num = month_map.get(m_abbr[:3])
                    if not m_num:
                        continue

                    dt_obj = datetime(ano, m_num, int(dia))
                    if dt_obj > DATA_HOJE:
                        continue

                    data_rows.append({
                        'Data': dt_obj,
                        'Cobre': clean_val(tds[1].text, 'metal'),
                        'Zinco': clean_val(tds[2].text, 'metal'),
                        'Chumbo': clean_val(tds[4].text, 'metal'),
                        'Niquel': clean_val(tds[6].text, 'metal'),
                        'Dolar': clean_val(tds[7].text, 'moeda')
                    })
            except:
                continue

    df = pd.DataFrame(data_rows)
    if not df.empty:
        df = df.drop_duplicates(subset='Data').set_index('Data')
    return df

@st.cache_data(ttl=3600)
def get_euro(data_inicio_str):
    try:
        euro_df = yf.download(
            "EURBRL=X",
            start=data_inicio_str,
            progress=False
        )['Close'].reset_index()

        euro_df.columns = ['Data', 'Euro']
        euro_df['Data'] = pd.to_datetime(euro_df['Data']).dt.tz_localize(None)
        return euro_df.set_index('Data')
    except:
        return pd.DataFrame()

def processar_planilha(uploaded_file):
    wb = load_workbook(uploaded_file)

    if NOME_ABA not in wb.sheetnames:
        return None, f"Erro: Aba '{NOME_ABA}' não encontrada no arquivo."

    ws = wb[NOME_ABA]

    # 1. Descobrir último ano preenchido
    ultimo_ano = 2024

    for row in ws.iter_rows(min_row=3, values_only=True):
        data_val, dolar_val = row[0], row[1]
        dt = None

        if isinstance(data_val, datetime):
            dt = data_val
        elif isinstance(data_val, str):
            try:
                dt = datetime.strptime(data_val.split()[0], "%d/%m/%Y")
            except:
                pass

        if dt and dt <= DATA_HOJE:
            if dolar_val not in [None, ""]:
                ultimo_ano = dt.year
            else:
                ultimo_ano = max(2024, dt.year)
                break

    start_year = max(2024, ultimo_ano)

    # 2. Coleta de dados
    df_site = get_shock_metais(start_year)
    df_euro = get_euro(f"{start_year}-01-01")

    df_completo = df_site.copy()
    if not df_euro.empty:
        df_completo = df_completo.join(df_euro, how='left')
        df_completo['Euro'] = df_completo['Euro'].ffill()  # 🔑 CORREÇÃO PRINCIPAL

    # 3. Preenchimento
    ups = 0

    for row_idx in range(3, 15000):
        cell_data = ws.cell(row=row_idx, column=1).value

        # Fim do arquivo
        if cell_data is None:
            vazios = sum(
                ws.cell(row=row_idx + k, column=1).value is None
                for k in range(1, 10)
            )
            if vazios == 9:
                break
            continue

        # Normaliza data
        dt = None
        if isinstance(cell_data, datetime):
            dt = cell_data.replace(hour=0, minute=0, second=0, microsecond=0)
        else:
            try:
                s = str(cell_data).split()[0]
                if '/' in s:
                    dt = datetime.strptime(s, "%d/%m/%Y")
                elif '-' in s:
                    dt = datetime.strptime(s, "%Y-%m-%d")
            except:
                continue

        if not dt or dt > DATA_HOJE:
            continue

        # Fim de semana
        if dt.weekday() in [5, 6]:
            for col in [2, 4, 6, 8, 10, 12]:
                c = ws.cell(row=row_idx, column=col)
                if c.value in [None, ""]:
                    c.value = "FIM DE SEMANA"
            ups += 1
            continue

        # Dias úteis
        if not df_completo.empty and dt in df_completo.index:
            dados = df_completo.loc[dt]
            mapeamento = {
                2: dados.get('Dolar'),
                4: dados.get('Euro'),
                6: dados.get('Cobre'),
                8: dados.get('Zinco'),
                10: dados.get('Niquel'),
                12: dados.get('Chumbo')
            }

            linha_mexida = False
            for col, val in mapeamento.items():
                c = ws.cell(row=row_idx, column=col)
                if pd.notna(val) and c.value in [None, "", 0]:
                    c.value = val
                    linha_mexida = True

            if linha_mexida:
                ups += 1

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return output, f"Sucesso! {ups} linhas foram atualizadas."

# --- INTERFACE STREAMLIT ---

st.set_page_config(page_title="Atualizador LME", page_icon="📈")

st.title("📈 Atualizador de Cotações LME")
st.markdown("""
Faça o upload da planilha **cotacao_bcb_lme.xlsx**.
O sistema irá:
1. Identificar datas faltantes.
2. Baixar dados da Shock Metais e Yahoo Finance.
3. Preencher lacunas (Sáb/Dom serão marcados como **FIM DE SEMANA**).
""")

uploaded_file = st.file_uploader("Escolha o arquivo Excel", type=["xlsx"])

if uploaded_file and st.button("🚀 Processar e Atualizar", type="primary"):
    with st.status("Processando...", expanded=True):
        processed_data, msg = processar_planilha(uploaded_file)

        if processed_data:
            st.success(msg)
            st.download_button(
                label="📥 Baixar Planilha Pronta",
                data=processed_data,
                file_name=f"cotacao_bcb_lme_atualizada_{datetime.now():%Y-%m-%d}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
        else:
            st.error(msg)
